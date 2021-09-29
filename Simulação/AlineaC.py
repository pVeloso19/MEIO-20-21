import numpy as np
import random as rand
from openpyxl import Workbook
from openpyxl.styles import Font

corVermelha = Font(color="FF0000")

t = 4

mediaEA, desvioPadraoEA = 459.9575, 1.024837 

mediaEB, desvioPadraoEB = 335.945, 0.45763 

numeroSemanas = 0

def isSemanaReverStock(semana):
    return (semana % t == 0)

def isEpocaAlta(semana):
    res = False
    
    while(semana>50):
        semana = semana-50

    if(semana >= 24 and semana <= 46):
        res = True
    
    return res

def get_LT():
    LT_List = []
    for i in range(28):
        LT_List.append(1)
    for i in range(60):
        LT_List.append(2)
    for i in range(12):
        LT_List.append(3)
    return rand.choice(LT_List)

def calcula(sEA,SEA,sEB,SEB,vezes):

    if(vezes==1):
        ws = wb.active
        ws.title = f'S{vezes}'
    else:
        ws = wb.create_sheet(title=f'S{vezes}')
    
    ws.cell(row=1, column=2, value='VALORES')
    ws.cell(row=3, column=2, value='Época Alta')
    ws.cell(row=4, column=1, value=f's={sEA}')
    ws.cell(row=5, column=1, value=f'S={SEA}')
    ws.cell(row=6, column=1, value=f't=4')
    ws.cell(row=8, column=2, value='Época Baixa')
    ws.cell(row=9, column=1, value=f's={sEB}')
    ws.cell(row=10, column=1, value=f'S={SEB}')
    ws.cell(row=11, column=1, value=f't=4')
    
    abastecimento = 0
    prazo = {}
    stock_Mao = sEA + int((1+sEA)/(2*rand.random()))
    stock_Final = stock_Mao
    stock_Inicial = 0
    encomenda = {}
    vendas = 0
    periodo = 0

    print('Periodo;Stock_Inicial;Prazo;Abastecimento;Vendas;Stock_Mao;Stock_Final;Encomenda')
    ws.cell(row=1, column=6, value='Periodo')
    ws.cell(row=1, column=7, value='Stock_Inicial')
    ws.cell(row=1, column=8, value='Prazo')
    ws.cell(row=1, column=9, value='Abastecimento')
    ws.cell(row=1, column=10, value='Vendas')
    ws.cell(row=1, column=11, value='Stock_Mao')
    ws.cell(row=1, column=12, value='Stock_Final')
    ws.cell(row=1, column=13, value='Encomenda')
    print(f'{periodo};{stock_Inicial};-;{abastecimento};{vendas};{stock_Mao};{stock_Final};0')
    ws.cell(row=2, column=6, value=periodo)
    ws.cell(row=2, column=7, value=stock_Inicial)
    ws.cell(row=2, column=8, value='-')
    ws.cell(row=2, column=9, value=abastecimento)
    ws.cell(row=2, column=10, value=vendas)
    ws.cell(row=2, column=11, value=stock_Mao)
    ws.cell(row=2, column=12, value=stock_Final)
    ws.cell(row=2, column=13, value=0)
    encomenda[periodo] = 0

    periodo = 1
    while(periodo <= numeroSemanas):

        abastecimento = 0

        stock_Inicial = stock_Final
        
        for (p,val) in prazo.items():
            if(val > 0):
                prazo[p] = val - 1
                if(prazo[p] == 0):
                    abastecimento = abastecimento + encomenda[p-1]
            
        if(encomenda[periodo-1] > 0):
            prazo[periodo] = get_LT()

        if(isEpocaAlta(periodo)):
            vendas = int(np.random.normal(mediaEA, desvioPadraoEA))  
        else:
            vendas = int(np.random.normal(mediaEB, desvioPadraoEB))  

        stock_Mao = encomenda[periodo-1] + stock_Mao - vendas
        stock_Final = stock_Inicial + abastecimento - vendas

        if(stock_Final<0):
            stock_Final = int(stock_Final*0.6)
        
        t_aux = isSemanaReverStock(periodo)

        if(isEpocaAlta(periodo+3)):
            if(t_aux and stock_Mao <= sEA):
                encomenda[periodo] = SEA - stock_Mao
            else:
                encomenda[periodo] = 0
        else:
            if(t_aux and stock_Mao <= sEB):
                encomenda[periodo] = SEB - stock_Mao
            else:
                encomenda[periodo] = 0

        print(f'{periodo};{stock_Inicial};', end='')
        ws.cell(row=periodo+2, column=6, value=periodo)
        ws.cell(row=periodo+2, column=7, value=stock_Inicial)
        ws.cell(row=periodo+2, column=9, value=abastecimento)
        ws.cell(row=periodo+2, column=10, value=vendas)
        ws.cell(row=periodo+2, column=11, value=stock_Mao)
        ws.cell(row=periodo+2, column=12, value=stock_Final)
        ws.cell(row=periodo+2, column=13, value=encomenda[periodo])
        single = True
        i = 0
        prazosStr = '-'
        for (p1,p2) in prazo.items():
            if(p2>=0):
                if(single):
                    print(f'{p2}', end='')
                    prazosStr = f'{p2}'
                    single = False
                    i += 1
                else:
                    print(f',{p2}', end='')
                    prazosStr += f',{p2}'
                    i += 1
                if(p2==0):
                    prazo[p1] = -1
        if(i==0):
            print('-', end='')
        ws.cell(row=periodo+2, column=8, value=prazosStr)
        print(f';{abastecimento};{vendas};{stock_Mao};{stock_Final};{encomenda[periodo]}', end='')
        print('')
        

        if(stock_Inicial < 0):
            a1 = ws[f'G{periodo+2}']
            a1.font = corVermelha

        periodo += 1

numSimulacoes = input("Quantas simulações a realizar: ")
anos = input("Número de anos a simular: ")
numeroSemanas = int(anos) * 50
gravar = input("Pretende gravar os dados simulados (S/N): ")

numSimulacoes = int(numSimulacoes)

for x in range(numSimulacoes):
    wb = Workbook()
    
    print('Simulação para (s,S)')
    print('')
    sEATemp = 2272
    SEATemp = 3352
    sEBTemp = 1411
    SEBTemp = 2448
    calcula(sEATemp, SEATemp, sEBTemp, SEBTemp, 1)
    print('')
    print('Simulação para (s+5%,S)')
    print('')
    sEATemp = 2386
    SEATemp = 3352
    sEBTemp = 1482
    SEBTemp = 2448
    calcula(sEATemp, SEATemp, sEBTemp, SEBTemp, 2)
    print('')
    print('Simulação para (s,S+5%)')
    print('')
    sEATemp = 2272
    SEATemp = 3520
    sEBTemp = 1411
    SEBTemp = 2360
    calcula(sEATemp, SEATemp, sEBTemp, SEBTemp, 3)
    print('')
    print('Simulação para (s-5%,S)')
    print('')
    sEATemp = 2158
    SEATemp = 3352
    sEBTemp = 1341
    SEBTemp = 2448
    calcula(sEATemp, SEATemp, sEBTemp, SEBTemp, 4)
    print('')
    print('Simulação para (s,S-5%)')
    print('')
    sEATemp = 2272
    SEATemp = 3184
    sEBTemp = 1411
    SEBTemp = 2136
    calcula(sEATemp, SEATemp, sEBTemp, SEBTemp, 5)
    print('')
    print('Simulação para (s+5%,S+5%)')
    print('')
    sEATemp = 2386
    SEATemp = 3520
    sEBTemp = 1482
    SEBTemp = 2570
    calcula(sEATemp, SEATemp, sEBTemp, SEBTemp, 6)
    print('')
    print('Simulação para (s-5%,S-5%)')
    print('')
    sEATemp = 2158
    SEATemp = 3352
    sEBTemp = 1411
    SEBTemp = 2326
    calcula(sEATemp, SEATemp, sEBTemp, SEBTemp, 7)

    if(gravar == 'S'):
        wb.save(f'simulacao{x}.xlsx')
    
    print('')
    print('')